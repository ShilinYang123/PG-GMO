#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel MCP Server
高效办公助手系统 - Excel文档处理MCP服务器
作者：雨俊
日期：2025-01-08
"""

import asyncio
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from fastmcp import FastMCP

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 创建MCP服务器实例
mcp = FastMCP("Excel Document Server")

class ExcelMCPServer:
    """Excel MCP服务器类"""
    
    def __init__(self):
        self.server_name = "Excel MCP Server"
        self.version = "1.0.0"
        self.workbooks = {}  # 存储打开的工作簿
        
    def initialize(self):
        """初始化Excel MCP服务器"""
        logger.info(f"初始化 {self.server_name} v{self.version}")
        
# 全局服务器实例
excel_server = ExcelMCPServer()

@mcp.tool()
async def create_workbook(file_path: str, sheet_names: Optional[List[str]] = None) -> Dict[str, Any]:
    """
    创建新的Excel工作簿
    
    Args:
        file_path: 文件保存路径
        sheet_names: 工作表名称列表，默认创建一个名为'Sheet1'的工作表
    
    Returns:
        包含操作结果的字典
    """
    try:
        # 创建新工作簿
        wb = openpyxl.Workbook()
        
        # 删除默认工作表
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # 创建指定的工作表
        if not sheet_names:
            sheet_names = ['Sheet1']
            
        for sheet_name in sheet_names:
            wb.create_sheet(title=sheet_name)
        
        # 保存工作簿
        wb.save(file_path)
        
        # 存储工作簿引用
        excel_server.workbooks[file_path] = wb
        
        return {
            "status": "success",
            "message": f"Excel工作簿已创建: {file_path}",
            "file_path": file_path,
            "sheet_names": sheet_names
        }
        
    except Exception as e:
        logger.error(f"创建Excel工作簿失败: {e}")
        return {
            "status": "error",
            "message": f"创建Excel工作簿失败: {str(e)}"
        }

@mcp.tool()
async def open_workbook(file_path: str) -> Dict[str, Any]:
    """
    打开现有的Excel工作簿
    
    Args:
        file_path: Excel文件路径
    
    Returns:
        包含工作簿信息的字典
    """
    try:
        if not Path(file_path).exists():
            return {
                "status": "error",
                "message": f"文件不存在: {file_path}"
            }
        
        # 打开工作簿
        wb = openpyxl.load_workbook(file_path)
        excel_server.workbooks[file_path] = wb
        
        # 获取工作表信息
        sheet_info = []
        for sheet in wb.worksheets:
            sheet_info.append({
                "name": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column
            })
        
        return {
            "status": "success",
            "message": f"Excel工作簿已打开: {file_path}",
            "file_path": file_path,
            "sheets": sheet_info
        }
        
    except Exception as e:
        logger.error(f"打开Excel工作簿失败: {e}")
        return {
            "status": "error",
            "message": f"打开Excel工作簿失败: {str(e)}"
        }

@mcp.tool()
async def write_data(file_path: str, sheet_name: str, data: List[List[Any]], 
                    start_row: int = 1, start_col: int = 1) -> Dict[str, Any]:
    """
    向Excel工作表写入数据
    
    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        data: 要写入的数据（二维列表）
        start_row: 起始行号（从1开始）
        start_col: 起始列号（从1开始）
    
    Returns:
        包含操作结果的字典
    """
    try:
        # 获取或打开工作簿
        if file_path not in excel_server.workbooks:
            await open_workbook(file_path)
        
        wb = excel_server.workbooks[file_path]
        
        # 获取工作表
        if sheet_name not in [sheet.title for sheet in wb.worksheets]:
            ws = wb.create_sheet(title=sheet_name)
        else:
            ws = wb[sheet_name]
        
        # 写入数据
        for row_idx, row_data in enumerate(data):
            for col_idx, cell_value in enumerate(row_data):
                ws.cell(row=start_row + row_idx, 
                       column=start_col + col_idx, 
                       value=cell_value)
        
        # 保存工作簿
        wb.save(file_path)
        
        return {
            "status": "success",
            "message": f"数据已写入工作表 '{sheet_name}'",
            "rows_written": len(data),
            "cols_written": len(data[0]) if data else 0
        }
        
    except Exception as e:
        logger.error(f"写入数据失败: {e}")
        return {
            "status": "error",
            "message": f"写入数据失败: {str(e)}"
        }

@mcp.tool()
async def read_data(file_path: str, sheet_name: str, 
                   start_row: int = 1, end_row: Optional[int] = None,
                   start_col: int = 1, end_col: Optional[int] = None) -> Dict[str, Any]:
    """
    从Excel工作表读取数据
    
    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        start_row: 起始行号
        end_row: 结束行号（None表示到最后一行）
        start_col: 起始列号
        end_col: 结束列号（None表示到最后一列）
    
    Returns:
        包含读取数据的字典
    """
    try:
        # 获取或打开工作簿
        if file_path not in excel_server.workbooks:
            await open_workbook(file_path)
        
        wb = excel_server.workbooks[file_path]
        ws = wb[sheet_name]
        
        # 确定读取范围
        if end_row is None:
            end_row = ws.max_row
        if end_col is None:
            end_col = ws.max_column
        
        # 读取数据
        data = []
        for row in ws.iter_rows(min_row=start_row, max_row=end_row,
                               min_col=start_col, max_col=end_col,
                               values_only=True):
            data.append(list(row))
        
        return {
            "status": "success",
            "data": data,
            "rows_read": len(data),
            "cols_read": len(data[0]) if data else 0
        }
        
    except Exception as e:
        logger.error(f"读取数据失败: {e}")
        return {
            "status": "error",
            "message": f"读取数据失败: {str(e)}"
        }

@mcp.tool()
async def format_cells(file_path: str, sheet_name: str, cell_range: str,
                      font_name: Optional[str] = None, font_size: Optional[int] = None,
                      bold: Optional[bool] = None, italic: Optional[bool] = None,
                      bg_color: Optional[str] = None, font_color: Optional[str] = None,
                      alignment: Optional[str] = None) -> Dict[str, Any]:
    """
    格式化Excel单元格
    
    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        cell_range: 单元格范围（如'A1:C3'）
        font_name: 字体名称
        font_size: 字体大小
        bold: 是否加粗
        italic: 是否斜体
        bg_color: 背景颜色（十六进制，如'FFFF00'）
        font_color: 字体颜色（十六进制）
        alignment: 对齐方式（'left', 'center', 'right'）
    
    Returns:
        包含操作结果的字典
    """
    try:
        # 获取或打开工作簿
        if file_path not in excel_server.workbooks:
            await open_workbook(file_path)
        
        wb = excel_server.workbooks[file_path]
        ws = wb[sheet_name]
        
        # 创建样式对象
        font = Font(name=font_name, size=font_size, bold=bold, italic=italic, color=font_color)
        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid') if bg_color else None
        align = Alignment(horizontal=alignment) if alignment else None
        
        # 应用格式到指定范围
        for row in ws[cell_range]:
            for cell in row:
                if font_name or font_size or bold is not None or italic is not None or font_color:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if align:
                    cell.alignment = align
        
        # 保存工作簿
        wb.save(file_path)
        
        return {
            "status": "success",
            "message": f"单元格格式已应用到范围 {cell_range}"
        }
        
    except Exception as e:
        logger.error(f"格式化单元格失败: {e}")
        return {
            "status": "error",
            "message": f"格式化单元格失败: {str(e)}"
        }

@mcp.tool()
async def create_chart(file_path: str, sheet_name: str, chart_type: str,
                      data_range: str, title: Optional[str] = None,
                      position: str = "E5") -> Dict[str, Any]:
    """
    在Excel工作表中创建图表
    
    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        chart_type: 图表类型（'bar', 'line', 'pie'）
        data_range: 数据范围（如'A1:B10'）
        title: 图表标题
        position: 图表位置（如'E5'）
    
    Returns:
        包含操作结果的字典
    """
    try:
        # 获取或打开工作簿
        if file_path not in excel_server.workbooks:
            await open_workbook(file_path)
        
        wb = excel_server.workbooks[file_path]
        ws = wb[sheet_name]
        
        # 创建数据引用
        data = Reference(ws, range_string=data_range)
        
        # 根据类型创建图表
        if chart_type.lower() == 'bar':
            chart = BarChart()
        elif chart_type.lower() == 'line':
            chart = LineChart()
        elif chart_type.lower() == 'pie':
            chart = PieChart()
        else:
            return {
                "status": "error",
                "message": f"不支持的图表类型: {chart_type}"
            }
        
        # 设置图表属性
        chart.add_data(data, titles_from_data=True)
        if title:
            chart.title = title
        
        # 添加图表到工作表
        ws.add_chart(chart, position)
        
        # 保存工作簿
        wb.save(file_path)
        
        return {
            "status": "success",
            "message": f"{chart_type}图表已创建在位置 {position}"
        }
        
    except Exception as e:
        logger.error(f"创建图表失败: {e}")
        return {
            "status": "error",
            "message": f"创建图表失败: {str(e)}"
        }

@mcp.tool()
async def get_workbook_info(file_path: str) -> Dict[str, Any]:
    """
    获取Excel工作簿信息
    
    Args:
        file_path: Excel文件路径
    
    Returns:
        包含工作簿信息的字典
    """
    try:
        # 获取或打开工作簿
        if file_path not in excel_server.workbooks:
            await open_workbook(file_path)
        
        wb = excel_server.workbooks[file_path]
        
        # 收集工作簿信息
        info = {
            "file_path": file_path,
            "sheet_count": len(wb.worksheets),
            "sheets": []
        }
        
        for sheet in wb.worksheets:
            sheet_info = {
                "name": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "has_data": sheet.max_row > 1 or sheet.max_column > 1
            }
            info["sheets"].append(sheet_info)
        
        return {
            "status": "success",
            "info": info
        }
        
    except Exception as e:
        logger.error(f"获取工作簿信息失败: {e}")
        return {
            "status": "error",
            "message": f"获取工作簿信息失败: {str(e)}"
        }

def run_server():
    """启动Excel MCP服务器"""
    excel_server.initialize()
    mcp.run()

if __name__ == "__main__":
    run_server()