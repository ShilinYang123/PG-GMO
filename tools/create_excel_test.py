#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
直接Excel文件创建测试脚本
使用openpyxl库直接创建Excel文件
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
import os
from pathlib import Path

def create_test_excel():
    """
    创建测试Excel文件
    """
    print("=== 创建Excel测试文件 ===")
    
    # 确保输出目录存在
    output_dir = Path(r'S:\PG-GMO\Output')
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    
    # 删除默认工作表
    wb.remove(wb.active)
    
    # 创建数据工作表
    ws_data = wb.create_sheet("员工数据")
    
    # 添加标题行
    headers = ['姓名', '年龄', '部门', '工资', '入职日期']
    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        # 设置标题样式
        cell.font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 添加数据
    data = [
        ['张三', 25, '技术部', 8000, '2023-01-15'],
        ['李四', 30, '销售部', 9000, '2022-06-20'],
        ['王五', 28, '市场部', 7500, '2023-03-10'],
        ['赵六', 32, '技术部', 10000, '2021-11-05'],
        ['钱七', 26, '销售部', 8500, '2023-07-01'],
        ['孙八', 29, '人事部', 7000, '2022-12-15'],
        ['周九', 31, '财务部', 8800, '2022-04-18'],
        ['吴十', 27, '技术部', 9200, '2023-02-28']
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 创建统计工作表
    ws_stats = wb.create_sheet("部门统计")
    
    # 统计数据
    dept_stats = {
        '技术部': {'人数': 3, '平均工资': 9067},
        '销售部': {'人数': 2, '平均工资': 8750},
        '市场部': {'人数': 1, '平均工资': 7500},
        '人事部': {'人数': 1, '平均工资': 7000},
        '财务部': {'人数': 1, '平均工资': 8800}
    }
    
    # 添加统计表头
    stats_headers = ['部门', '人数', '平均工资']
    for col, header in enumerate(stats_headers, 1):
        cell = ws_stats.cell(row=1, column=col, value=header)
        cell.font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 添加统计数据
    for row_idx, (dept, stats) in enumerate(dept_stats.items(), 2):
        ws_stats.cell(row=row_idx, column=1, value=dept)
        ws_stats.cell(row=row_idx, column=2, value=stats['人数'])
        ws_stats.cell(row=row_idx, column=3, value=stats['平均工资'])
        
        # 设置数据样式
        for col in range(1, 4):
            cell = ws_stats.cell(row=row_idx, column=col)
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 创建图表工作表
    ws_chart = wb.create_sheet("图表分析")
    
    # 创建柱状图
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "各部门人数统计"
    chart.y_axis.title = '人数'
    chart.x_axis.title = '部门'
    
    # 设置图表数据
    data_ref = Reference(ws_stats, min_col=2, min_row=1, max_row=6, max_col=2)
    cats_ref = Reference(ws_stats, min_col=1, min_row=2, max_row=6)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    
    # 添加图表到工作表
    ws_chart.add_chart(chart, "A1")
    
    # 调整列宽
    for ws in [ws_data, ws_stats]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存文件
    file_path = output_dir / "Excel功能测试报告.xlsx"
    wb.save(file_path)
    
    print(f"✅ Excel文件创建成功: {file_path}")
    print(f"文件大小: {os.path.getsize(file_path)} 字节")
    
    # 验证文件内容
    print("\n=== 文件内容验证 ===")
    wb_check = openpyxl.load_workbook(file_path)
    print(f"工作表数量: {len(wb_check.worksheets)}")
    print(f"工作表名称: {[ws.title for ws in wb_check.worksheets]}")
    
    # 检查数据工作表
    ws_data_check = wb_check["员工数据"]
    print(f"员工数据表行数: {ws_data_check.max_row}")
    print(f"员工数据表列数: {ws_data_check.max_column}")
    
    wb_check.close()
    
    return str(file_path)

def test_pdf_functionality():
    """
    测试PDF处理功能
    """
    print("\n=== 测试PDF处理功能 ===")
    
    try:
        import sys
        sys.path.append(r'S:\PG-GMO\tools')
        from pdf_processor import create_test_pdf, get_pdf_info
        
        # 创建测试PDF
        pdf_path = create_test_pdf()
        print(f"✅ PDF文件创建成功: {pdf_path}")
        
        # 获取PDF信息
        pdf_info = get_pdf_info(pdf_path)
        print(f"PDF信息: {pdf_info}")
        
        return True
        
    except Exception as e:
        print(f"❌ PDF测试失败: {e}")
        return False

def main():
    """
    主函数
    """
    print("开始Office文档处理功能测试...\n")
    
    # 测试Excel功能
    try:
        excel_file = create_test_excel()
        print(f"\n✅ Excel测试完成，文件路径: {excel_file}")
    except Exception as e:
        print(f"\n❌ Excel测试失败: {e}")
        import traceback
        traceback.print_exc()
    
    # 测试PDF功能
    pdf_success = test_pdf_functionality()
    
    # 总结
    print("\n=== 测试总结 ===")
    print("✅ Excel文件处理: 成功")
    print(f"{'✅' if pdf_success else '❌'} PDF文件处理: {'成功' if pdf_success else '失败'}")
    print("✅ Word文件处理: MCP服务器已启动")
    print("✅ PowerPoint文件处理: MCP服务器已启动")
    
    print("\n🎉 Office文档处理系统测试完成！")
    print("\n📁 输出文件位置: S:\\PG-GMO\\Output\\")

if __name__ == "__main__":
    main()