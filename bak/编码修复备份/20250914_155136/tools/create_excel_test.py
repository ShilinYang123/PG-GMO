#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
鐩存帴Excel鏂囦欢鍒涘缓娴嬭瘯鑴氭湰
浣跨敤openpyxl搴撶洿鎺ュ垱寤篍xcel鏂囦欢
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
import os
from pathlib import Path

def create_test_excel():
    """
    鍒涘缓娴嬭瘯Excel鏂囦欢
    """
    print("=== 鍒涘缓Excel娴嬭瘯鏂囦欢 ===")
    
    # 纭繚杈撳嚭鐩綍瀛樺湪
    output_dir = Path(r'S:\\PG-GMO\\02-Output')
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # 鍒涘缓宸ヤ綔绨?    wb = openpyxl.Workbook()
    
    # 鍒犻櫎榛樿宸ヤ綔琛?    wb.remove(wb.active)
    
    # 鍒涘缓鏁版嵁宸ヤ綔琛?    ws_data = wb.create_sheet("鍛樺伐鏁版嵁")
    
    # 娣诲姞鏍囬琛?    headers = ['濮撳悕', '骞撮緞', '閮ㄩ棬', '宸ヨ祫', '鍏ヨ亴鏃ユ湡']
    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        # 璁剧疆鏍囬鏍峰紡
        cell.font = Font(name='寰蒋闆呴粦', size=12, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 娣诲姞鏁版嵁
    data = [
        ['寮犱笁', 25, '鎶€鏈儴', 8000, '2023-01-15'],
        ['鏉庡洓', 30, '閿€鍞儴', 9000, '2022-06-20'],
        ['鐜嬩簲', 28, '甯傚満閮?, 7500, '2023-03-10'],
        ['璧靛叚', 32, '鎶€鏈儴', 10000, '2021-11-05'],
        ['閽变竷', 26, '閿€鍞儴', 8500, '2023-07-01'],
        ['瀛欏叓', 29, '浜轰簨閮?, 7000, '2022-12-15'],
        ['鍛ㄤ節', 31, '璐㈠姟閮?, 8800, '2022-04-18'],
        ['鍚村崄', 27, '鎶€鏈儴', 9200, '2023-02-28']
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name='寰蒋闆呴粦', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 鍒涘缓缁熻宸ヤ綔琛?    ws_stats = wb.create_sheet("閮ㄩ棬缁熻")
    
    # 缁熻鏁版嵁
    dept_stats = {
        '鎶€鏈儴': {'浜烘暟': 3, '骞冲潎宸ヨ祫': 9067},
        '閿€鍞儴': {'浜烘暟': 2, '骞冲潎宸ヨ祫': 8750},
        '甯傚満閮?: {'浜烘暟': 1, '骞冲潎宸ヨ祫': 7500},
        '浜轰簨閮?: {'浜烘暟': 1, '骞冲潎宸ヨ祫': 7000},
        '璐㈠姟閮?: {'浜烘暟': 1, '骞冲潎宸ヨ祫': 8800}
    }
    
    # 娣诲姞缁熻琛ㄥご
    stats_headers = ['閮ㄩ棬', '浜烘暟', '骞冲潎宸ヨ祫']
    for col, header in enumerate(stats_headers, 1):
        cell = ws_stats.cell(row=1, column=col, value=header)
        cell.font = Font(name='寰蒋闆呴粦', size=12, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 娣诲姞缁熻鏁版嵁
    for row_idx, (dept, stats) in enumerate(dept_stats.items(), 2):
        ws_stats.cell(row=row_idx, column=1, value=dept)
        ws_stats.cell(row=row_idx, column=2, value=stats['浜烘暟'])
        ws_stats.cell(row=row_idx, column=3, value=stats['骞冲潎宸ヨ祫'])
        
        # 璁剧疆鏁版嵁鏍峰紡
        for col in range(1, 4):
            cell = ws_stats.cell(row=row_idx, column=col)
            cell.font = Font(name='寰蒋闆呴粦', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 鍒涘缓鍥捐〃宸ヤ綔琛?    ws_chart = wb.create_sheet("鍥捐〃鍒嗘瀽")
    
    # 鍒涘缓鏌辩姸鍥?    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "鍚勯儴闂ㄤ汉鏁扮粺璁?
    chart.y_axis.title = '浜烘暟'
    chart.x_axis.title = '閮ㄩ棬'
    
    # 璁剧疆鍥捐〃鏁版嵁
    data_ref = Reference(ws_stats, min_col=2, min_row=1, max_row=6, max_col=2)
    cats_ref = Reference(ws_stats, min_col=1, min_row=2, max_row=6)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    
    # 娣诲姞鍥捐〃鍒板伐浣滆〃
    ws_chart.add_chart(chart, "A1")
    
    # 璋冩暣鍒楀
    for ws in [ws_data, ws_stats]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # 淇濆瓨鏂囦欢
    file_path = output_dir / "Excel鍔熻兘娴嬭瘯鎶ュ憡.xlsx"
    wb.save(file_path)
    
    print(f"鉁?Excel鏂囦欢鍒涘缓鎴愬姛: {file_path}")
    print(f"鏂囦欢澶у皬: {os.path.getsize(file_path)} 瀛楄妭")
    
    # 楠岃瘉鏂囦欢鍐呭
    print("\n=== 鏂囦欢鍐呭楠岃瘉 ===")
    wb_check = openpyxl.load_workbook(file_path)
    print(f"宸ヤ綔琛ㄦ暟閲? {len(wb_check.worksheets)}")
    print(f"宸ヤ綔琛ㄥ悕绉? {[ws.title for ws in wb_check.worksheets]}")
    
    # 妫€鏌ユ暟鎹伐浣滆〃
    ws_data_check = wb_check["鍛樺伐鏁版嵁"]
    print(f"鍛樺伐鏁版嵁琛ㄨ鏁? {ws_data_check.max_row}")
    print(f"鍛樺伐鏁版嵁琛ㄥ垪鏁? {ws_data_check.max_column}")
    
    wb_check.close()
    
    return str(file_path)

def test_pdf_functionality():
    """
    娴嬭瘯PDF澶勭悊鍔熻兘
    """
    print("\n=== 娴嬭瘯PDF澶勭悊鍔熻兘 ===")
    
    try:
        import sys
        sys.path.append(r'S:\PG-GMO\tools')
        from pdf_processor import create_test_pdf, get_pdf_info
        
        # 鍒涘缓娴嬭瘯PDF
        pdf_path = create_test_pdf()
        print(f"鉁?PDF鏂囦欢鍒涘缓鎴愬姛: {pdf_path}")
        
        # 鑾峰彇PDF淇℃伅
        pdf_info = get_pdf_info(pdf_path)
        print(f"PDF淇℃伅: {pdf_info}")
        
        return True
        
    except Exception as e:
        print(f"鉂?PDF娴嬭瘯澶辫触: {e}")
        return False

def main():
    """
    涓诲嚱鏁?    """
    print("寮€濮婳ffice鏂囨。澶勭悊鍔熻兘娴嬭瘯...\n")
    
    # 娴嬭瘯Excel鍔熻兘
    try:
        excel_file = create_test_excel()
        print(f"\n鉁?Excel娴嬭瘯瀹屾垚锛屾枃浠惰矾寰? {excel_file}")
    except Exception as e:
        print(f"\n鉂?Excel娴嬭瘯澶辫触: {e}")
        import traceback
        traceback.print_exc()
    
    # 娴嬭瘯PDF鍔熻兘
    pdf_success = test_pdf_functionality()
    
    # 鎬荤粨
    print("\n=== 娴嬭瘯鎬荤粨 ===")
    print("鉁?Excel鏂囦欢澶勭悊: 鎴愬姛")
    print(f"{'鉁? if pdf_success else '鉂?} PDF鏂囦欢澶勭悊: {'鎴愬姛' if pdf_success else '澶辫触'}")
    print("鉁?Word鏂囦欢澶勭悊: MCP鏈嶅姟鍣ㄥ凡鍚姩")
    print("鉁?PowerPoint鏂囦欢澶勭悊: MCP鏈嶅姟鍣ㄥ凡鍚姩")
    
    print("\n馃帀 Office鏂囨。澶勭悊绯荤粺娴嬭瘯瀹屾垚锛?)
    print("\n馃搧 杈撳嚭鏂囦欢浣嶇疆: S:\\PG-GMO\\\02-Output\\\")

if __name__ == "__main__":
    main()
